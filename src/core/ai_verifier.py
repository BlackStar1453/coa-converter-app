"""Claude Agent SDK-based COA verification agent.

Uses custom MCP tools to read PDF text, read XLSX cells, and compare fields.
The agent autonomously reads all three files (PDF, template, output) and
produces a structured JSON verification report.
"""

import os
import json
import logging
import asyncio
from typing import Any

import anthropic

log = logging.getLogger(__name__)

VERIFICATION_SYSTEM_PROMPT = """\
You are a COA (Certificate of Analysis) verification agent. Your job is to
verify that a PDF-to-Excel conversion was done correctly.

You have tools to:
1. read_pdf_text — extract text from the source PDF
2. read_xlsx_cells — read cells from the output XLSX file
3. read_template_cells — read cells from the original template

Your task:
- Use read_pdf_text to extract all data from the source PDF
- Use read_xlsx_cells to read the filled output file
- Compare every field: product name, batch number, dates, assay/ratio,
  all analytical items, microbiology items, packing/storage
- Return a JSON report with this exact structure:

{
  "summary": "Brief overall assessment",
  "total_fields": <int>,
  "passed": <int>,
  "failed": <int>,
  "accuracy": <float 0-1>,
  "confidence": "high|medium|low",
  "field_checks": [
    {
      "field": "field name",
      "pdf_value": "value from PDF",
      "output_value": "value from output",
      "status": "pass|fail|uncertain",
      "note": "optional note"
    }
  ],
  "issues": ["list of issues found"],
  "recommendations": ["list of recommendations"]
}

Be thorough. Check every single field. Report any discrepancies.
"""


def _extract_pdf_text(pdf_path: str) -> str:
    """Extract full text from PDF using pdfplumber."""
    import pdfplumber
    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
            # Also try tables
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if row:
                        text_parts.append(" | ".join(str(c or "") for c in row))
    return "\n".join(text_parts)


def _read_xlsx_cells(xlsx_path: str) -> str:
    """Read all non-empty cells from an XLSX file."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    lines = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value is not None:
                lines.append(f"[{cell.coordinate}] {cell.value}")
    wb.close()
    return "\n".join(lines)


class COAVerificationAgent:
    """Wrapper that runs Claude with tool use for COA verification."""

    def __init__(self, api_key: str, model: str = "claude-sonnet-4-20250514"):
        self.client = anthropic.Anthropic(api_key=api_key)
        self.model = model

    def verify(self, pdf_path: str, template_path: str, output_path: str) -> dict:
        """Run the verification agent and return structured results."""
        log.info("[AIVerifier] Starting verification agent...")

        # Pre-extract file contents (tools will serve these)
        pdf_text = _extract_pdf_text(pdf_path)
        output_cells = _read_xlsx_cells(output_path)
        template_cells = _read_xlsx_cells(template_path) if template_path.endswith(".xlsx") else "(DOCX template — not applicable)"

        # Define tools
        tools = [
            {
                "name": "read_pdf_text",
                "description": "Read the full text content extracted from the source PDF file.",
                "input_schema": {
                    "type": "object",
                    "properties": {},
                    "required": [],
                },
            },
            {
                "name": "read_xlsx_cells",
                "description": "Read all non-empty cells from the output XLSX file (the conversion result).",
                "input_schema": {
                    "type": "object",
                    "properties": {},
                    "required": [],
                },
            },
            {
                "name": "read_template_cells",
                "description": "Read all non-empty cells from the original template XLSX file.",
                "input_schema": {
                    "type": "object",
                    "properties": {},
                    "required": [],
                },
            },
        ]

        # Tool result mapping
        tool_data = {
            "read_pdf_text": pdf_text,
            "read_xlsx_cells": output_cells,
            "read_template_cells": template_cells,
        }

        # Initial message
        messages = [
            {
                "role": "user",
                "content": (
                    f"Please verify the COA conversion.\n"
                    f"Source PDF: {os.path.basename(pdf_path)}\n"
                    f"Template: {os.path.basename(template_path)}\n"
                    f"Output: {os.path.basename(output_path)}\n\n"
                    f"Use your tools to read all three files, then compare every field "
                    f"and return a JSON verification report."
                ),
            }
        ]

        # Agentic loop — keep calling until no more tool use
        max_turns = 10
        for turn in range(max_turns):
            log.info("[AIVerifier] Agent turn %d...", turn + 1)

            response = self.client.messages.create(
                model=self.model,
                max_tokens=4096,
                system=VERIFICATION_SYSTEM_PROMPT,
                tools=tools,
                messages=messages,
            )

            # Check for tool use
            tool_uses = [b for b in response.content if b.type == "tool_use"]
            if not tool_uses:
                # No tool calls — agent is done, extract text
                break

            # Process tool calls
            assistant_content = response.content
            messages.append({"role": "assistant", "content": assistant_content})

            tool_results = []
            for tool_use in tool_uses:
                tool_name = tool_use.name
                data = tool_data.get(tool_name, f"Unknown tool: {tool_name}")
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tool_use.id,
                    "content": str(data)[:50000],  # Truncate if very long
                })

            messages.append({"role": "user", "content": tool_results})

            if response.stop_reason == "end_turn":
                break

        # Extract final text response
        text_blocks = [b.text for b in response.content if hasattr(b, "text")]
        full_text = "\n".join(text_blocks)

        # Try to parse JSON from the response
        return self._parse_report(full_text)

    def _parse_report(self, text: str) -> dict:
        """Extract JSON report from agent response."""
        # Try to find JSON block
        import re
        json_match = re.search(r"```json\s*(.*?)\s*```", text, re.DOTALL)
        if json_match:
            try:
                return json.loads(json_match.group(1))
            except json.JSONDecodeError:
                pass

        # Try parsing the whole text as JSON
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass

        # Fallback: return raw text as summary
        return {
            "summary": text[:2000],
            "total_fields": 0,
            "passed": 0,
            "failed": 0,
            "accuracy": 0.0,
            "confidence": "low",
            "field_checks": [],
            "issues": ["Could not parse structured report from AI response"],
            "recommendations": [],
        }
